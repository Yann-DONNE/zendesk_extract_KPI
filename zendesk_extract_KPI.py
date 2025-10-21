import requests
import json
from requests.auth import HTTPBasicAuth
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm
import time
import os

# ----------------------------- Entête console -----------------------------
def print_header_corporate():
    BLUE = "\033[94m"
    GREEN = "\033[92m"
    YELLOW = "\033[93m"
    RESET = "\033[0m"
    BOLD = "\033[1m"

    header_lines = [
        f"{BLUE}{BOLD}=================================================={RESET}",
        f"{GREEN}{BOLD}          SCRIPT D'EXTRACTION KPI ZENDESK        {RESET}",
        f"{BLUE}{BOLD}=================================================={RESET}",
        f"{BOLD}Auteur :{RESET} Yann Donne",
        f"{BOLD}Date   :{RESET} 2025",
        f"{BOLD}Version:{RESET} 3.1 (Correction satisfaction mensuelle)",
        f"{YELLOW}--------------------------------------------------{RESET}"
    ]

    for line in header_lines:
        print(line)

# Affichage de l'entête au lancement
print_header_corporate()

# ----------------------------- Chrono début -----------------------------
start_time = time.time()

# ----------------------------- Fonctions utilitaires -----------------------------
def load_config(file_path='config.json'):
    """Charge les paramètres de connexion depuis un fichier JSON."""
    if not os.path.exists(file_path):
        print(f"❌ Erreur: Le fichier de configuration '{file_path}' est introuvable.")
        print("Veuillez créer un fichier 'config.json' avec les clés SUBDOMAIN, EMAIL et API_TOKEN.")
        return None
    with open(file_path, 'r') as f:
        config = json.load(f)
    return config

def get_tickets_incremental(auth, subdomain, start_date, end_date):
    """Récupère les tickets de Zendesk via l'API incrémentale."""
    url = f"https://{subdomain}.zendesk.com/api/v2/incremental/tickets.json?start_time={int(start_date.timestamp())}"
    all_tickets = []
    print("🔍 Chargement des tickets...")
    while url:
        try:
            resp = requests.get(url, auth=auth, timeout=30)
            if resp.status_code == 429:
                retry_after = int(resp.headers.get("Retry-After", 10))
                print(f"⏳ Rate limit atteint -> pause {retry_after}s...")
                time.sleep(retry_after)
                continue
            if resp.status_code != 200:
                print(f"❌ Erreur API {resp.status_code}: {resp.text}")
                break

            data = resp.json()
            for t in data.get("tickets", []):
                created = datetime.strptime(t["created_at"], "%Y-%m-%dT%H:%M:%SZ")
                if start_date <= created <= end_date:
                    all_tickets.append(t)
            
            if data.get("end_of_stream"):
                break
            url = data.get("next_page")
        except Exception as e:
            print(f"❌ Exception pendant la récupération des tickets: {e}")
            break
    print(f"✔ {len(all_tickets)} tickets récupérés")
    return all_tickets

def get_ticket_metrics(auth, subdomain, ticket_id, retries=3):
    """Récupère les métriques d'un ticket avec une gestion des erreurs et des retries."""
    url = f"https://{subdomain}.zendesk.com/api/v2/tickets/{ticket_id}/metrics.json"
    try:
        response = requests.get(url, auth=auth, timeout=10)
        if response.status_code == 200:
            data = response.json().get("ticket_metric", {})
            first_reply_time = data.get("reply_time_in_minutes", {}).get("calendar")
            resolution_time = data.get("full_resolution_time_in_minutes", {}).get("calendar")
            return ticket_id, first_reply_time, resolution_time
        elif response.status_code == 429 and retries > 0:
            retry_after = int(response.headers.get("Retry-After", 5))
            time.sleep(retry_after)
            return get_ticket_metrics(auth, subdomain, ticket_id, retries - 1)
        else:
            return ticket_id, None, None
    except Exception:
        if retries > 0:
            time.sleep(2)
            return get_ticket_metrics(auth, subdomain, ticket_id, retries - 1)
        return ticket_id, None, None

def process_data(tickets, all_types):
    """Traite les données brutes des tickets pour calculer les KPI."""
    
    for t in tickets:
        if t.get("type") == "problem":
            t["type"] = "incident"

    tag_data = defaultdict(lambda: {"types": defaultdict(int), "total": 0})
    unique_tagged_by_type = {t: set() for t in all_types}
    unique_tagged_total = set()

    for ticket in tickets:
        ttype = ticket.get("type")
        tid = ticket.get("id")
        tags = ticket.get("tags", [])
        has_com = False
        for tag in tags:
            if tag.startswith("com") and ttype in all_types:
                tag_data[tag]["types"][ttype] += 1
                tag_data[tag]["total"] += 1
                has_com = True
        if has_com and ttype in all_types and tid is not None:
            unique_tagged_by_type[ttype].add(tid)
            unique_tagged_total.add(tid)
    
    def sort_com_tags(tags):
        def extract_number(tag):
            match = re.search(r'com(\d+)', tag)
            return int(match.group(1)) if match else float('inf')
        return sorted(tags, key=extract_number)
    com_tags_sorted = sort_com_tags([tag for tag in tag_data if tag.startswith("com")])

    delai_first_reply = {'0-1h': 0, '1-8h': 0, '8-24h': 0, '>24h': 0}
    delai_resolution = {'0-5h': 0, '5-24h': 0, '1-7j': 0, '7-30j': 0, '>30j': 0}
    
    ticket_lookup = {ticket['id']: ticket for ticket in tickets if 'id' in ticket}
    
    return {
        "tag_data": tag_data,
        "unique_tagged_by_type": unique_tagged_by_type,
        "unique_tagged_total": unique_tagged_total,
        "com_tags_sorted": com_tags_sorted,
        "ticket_lookup": ticket_lookup,
        "delai_first_reply": delai_first_reply,
        "delai_resolution": delai_resolution
    }

def collect_metrics(auth, subdomain, tickets_to_process, data_structs):
    """Collecte les métriques de réponse et de résolution en parallèle."""
    
    delai_first_reply = data_structs["delai_first_reply"]
    delai_resolution = data_structs["delai_resolution"]
    ticket_lookup = data_structs["ticket_lookup"]
    all_types = ["incident", "question", "task"]
    
    print("⏳ Récupération des métriques des tickets...")
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {executor.submit(get_ticket_metrics, auth, subdomain, tid): tid for tid in tickets_to_process}
        for future in tqdm(as_completed(futures), total=len(futures), desc="Traitement des métriques"):
            try:
                ticket_id, first_reply_time, resolution_time = future.result()
            except Exception:
                continue

            ticket_obj = ticket_lookup.get(ticket_id)
            if not ticket_obj:
                continue
            ttype = ticket_obj.get('type')
            
            if ttype in all_types:
                if first_reply_time is not None:
                    if first_reply_time <= 60: delai_first_reply['0-1h'] += 1
                    elif first_reply_time <= 480: delai_first_reply['1-8h'] += 1
                    elif first_reply_time <= 1440: delai_first_reply['8-24h'] += 1
                    else: delai_first_reply['>24h'] += 1
                
                if resolution_time is not None:
                    if resolution_time <= 300: delai_resolution['0-5h'] += 1
                    elif resolution_time <= 1440: delai_resolution['5-24h'] += 1
                    elif resolution_time <= 10080: delai_resolution['1-7j'] += 1
                    elif resolution_time <= 43200: delai_resolution['7-30j'] += 1
                    else: delai_resolution['>30j'] += 1

def generate_excel_report(data_structs, tickets):
    """Génère le fichier Excel avec les différents KPI."""
    
    wb = Workbook()
    ws_tags = wb.active
    ws_tags.title = "Tickets par Tags"
    all_types = ["incident", "question", "task"]

    def add_colored_header(ws, headers):
        fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Bleu
        font = Font(bold=True, color="FFFFFF")  # Texte blanc gras
        alignment = Alignment(horizontal="center", vertical="center")
        ws.append(headers)
        for col_num in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment

    # --- Onglet Tags ---
    header_tags = ["Tag"] + all_types + ["Total tickets"]
    add_colored_header(ws_tags, header_tags)
    for tag in data_structs["com_tags_sorted"]:
        data = data_structs["tag_data"][tag]
        row = [tag] + [data['types'].get(t, 0) for t in all_types] + [data['total']]
        ws_tags.append(row)

    ws_tags.append([])
    ws_tags.append(["--- Résumé cohérence ---"])
    summary_header = ["Indicateur"] + all_types + ["Total"]
    ws_tags.append(summary_header)
    
    tickets_par_type = defaultdict(int)
    for ticket in tickets:
        ttype = ticket.get('type') or 'inconnu'
        tickets_par_type[ttype] += 1
    total_tickets = sum([tickets_par_type[t] for t in all_types])
    
    unique_tagged_row = ["Tickets avec >=1 tag 'com'"] + [len(data_structs["unique_tagged_by_type"].get(t, set())) for t in all_types] + [len(data_structs["unique_tagged_total"])]
    ws_tags.append(unique_tagged_row)

    sans_tag_row = ["Tickets SANS tag 'com'"]
    sum_sans_tag = 0
    for t in all_types:
        sans = tickets_par_type.get(t, 0) - len(data_structs["unique_tagged_by_type"].get(t, set()))
        if sans < 0: sans = 0
        sans_tag_row.append(sans)
        sum_sans_tag += sans
    sans_tag_row.append(sum_sans_tag)
    ws_tags.append(sans_tag_row)

    total_row = ["Total tickets (par type)"] + [tickets_par_type.get(t, 0) for t in all_types] + [total_tickets]
    ws_tags.append(total_row)
    
    # --- Onglet Délai 1ère Prise ---
    ws_delai = wb.create_sheet(title="Délai 1ère Prise")
    add_colored_header(ws_delai, ["Délai", "% Tickets", "Nombre"])
    sum_cats_first = sum(data_structs["delai_first_reply"].values())
    for categorie, count in data_structs["delai_first_reply"].items():
        pct = round((count / total_tickets) * 100, 1) if total_tickets > 0 else 0
        ws_delai.append([categorie, f"{pct}%", count])
    sans_metric_first = total_tickets - sum_cats_first
    if sans_metric_first < 0: sans_metric_first = 0
    pct_sans_first = round((sans_metric_first / total_tickets) * 100, 1) if total_tickets > 0 else 0
    ws_delai.append(["Sans métrique", f"{pct_sans_first}%", sans_metric_first])

    # --- Onglet Délai Résolution ---
    ws_resol = wb.create_sheet(title="Délai Résolution Complète")
    add_colored_header(ws_resol, ["Délai", "% Tickets", "Nombre"])
    sum_cats_res = sum(data_structs["delai_resolution"].values())
    for categorie, count in data_structs["delai_resolution"].items():
        pct = round((count / total_tickets) * 100, 1) if total_tickets > 0 else 0
        ws_resol.append([categorie, f"{pct}%", count])
    sans_metric_res = total_tickets - sum_cats_res
    if sans_metric_res < 0: sans_metric_res = 0
    pct_sans_res = round((sans_metric_res / total_tickets) * 100, 1) if total_tickets > 0 else 0
    ws_resol.append(["Sans métrique", f"{pct_sans_res}%", sans_metric_res])

    # --- Onglet Satisfaction ---
    ws_satisfaction = wb.create_sheet(title="Satisfaction")
    add_colored_header(ws_satisfaction, ["Indicateur", "Valeur"])
    satisfaction_counts = {'good': 0, 'bad': 0}
    total_notes = 0
    for ticket in tickets:
        satisfaction = ticket.get('satisfaction_rating')
        if satisfaction and isinstance(satisfaction, dict):
            score = satisfaction.get('score')
            if score in ('good', 'bad'):
                satisfaction_counts[score] += 1
                total_notes += 1
    pct_satisfaction = round((satisfaction_counts['good'] / total_notes) * 100, 1) if total_notes > 0 else 0.0
    ws_satisfaction.append(["% Satisfaction Globale", f"{pct_satisfaction}%"])
    ws_satisfaction.append(["Nombre 'Good'", satisfaction_counts['good']])
    ws_satisfaction.append(["Nombre 'Bad'", satisfaction_counts['bad']])
    
    # --- Onglet Tickets par Type ---
    ws_type = wb.create_sheet(title="Tickets par Type")
    add_colored_header(ws_type, ["Type de ticket", "Nombre"])
    for ttype in all_types:
        count = tickets_par_type.get(ttype, 0)
        pct = round((count / total_tickets) * 100, 1) if total_tickets > 0 else 0
        ws_type.append([ttype, f"{count} ({pct}%)"])
    ws_type.append(["Total", total_tickets])
    
    # --- Onglet Tickets par Mois ---
    ws_mois = wb.create_sheet(title="Tickets par Mois")
    add_colored_header(ws_mois, ["Mois", "Incidents", "Questions", "Tasks", "Total tickets", "% Satisfaction", "Nb Avis"])
    mois_fr = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
               "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
    stats_mensuelles = defaultdict(lambda: {'incident': 0, 'question': 0, 'task': 0, 'total': 0, 'good': 0, 'bad': 0})

    for ticket in tickets:
        satisfaction = ticket.get('satisfaction_rating')
        if satisfaction and isinstance(satisfaction, dict) and satisfaction.get('created_at'):
            mois_cle = satisfaction['created_at'][:7]
        else:
            created_at = ticket.get('created_at')
            mois_cle = created_at[:7] if created_at else "inconnu"

        ttype = ticket.get('type')
        if ttype in all_types:
            stats_mensuelles[mois_cle][ttype] += 1
            stats_mensuelles[mois_cle]['total'] += 1

        if satisfaction and isinstance(satisfaction, dict):
            score = satisfaction.get('score')
            if score in ('good', 'bad'):
                stats_mensuelles[mois_cle][score] += 1

    mois_tries = sorted(stats_mensuelles.keys())
    align_center = Alignment(horizontal='center', vertical='center')
    for mois_cle in mois_tries:
        data = stats_mensuelles[mois_cle]
        annee, mois_num = mois_cle.split('-')
        mois_fr_str = mois_fr[int(mois_num)-1] + " " + annee
        total_avis = data['good'] + data['bad']
        pct_sat = (data['good'] / total_avis * 100) if total_avis > 0 else 0.0
        ligne = [mois_fr_str, data['incident'], data['question'], data['task'], data['total'], f"{round(pct_sat, 1)}%", total_avis]
        ws_mois.append(ligne)
    for row in ws_mois.iter_rows(min_row=2, max_row=ws_mois.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = align_center

    wb.save("tickets_par_tags_et_delais.xlsx")
    print("✔ Fichier Excel généré : tickets_par_tags_et_delais.xlsx")

# ----------------------------- Fonction principale -----------------------------
def main():
    config = load_config()
    if not config:
        return

    auth = HTTPBasicAuth(f'{config["EMAIL"]}/token', config["API_TOKEN"])
    all_types = ["incident", "question", "task"]

    # --- Dates d'extraction ---
    while True:
        try:
            date_input = input("Entrez la date de **début** (JJ/MM/YYYY) ou Entrée pour 01/01 : ")
            start_date = datetime.strptime(date_input, '%d/%m/%Y') if date_input.strip() else datetime(datetime.now().year, 1, 1)
            break
        except ValueError:
            print("❌ Format invalide. Veuillez réessayer.")

    while True:
        try:
            end_input = input("Entrez la date de **fin** (JJ/MM/YYYY) ou Entrée pour aujourd'hui : ")
            end_date = datetime.strptime(end_input, '%d/%m/%Y') if end_input.strip() else datetime.now()
            break
        except ValueError:
            print("❌ Format invalide. Veuillez réessayer.")
            
    print(f"📅 Extraction des tickets du {start_date.date()} au {end_date.date()}")

    tickets = get_tickets_incremental(auth, config["SUBDOMAIN"], start_date, end_date)
    if not tickets:
        print("Aucun ticket récupéré. Arrêt du script.")
        return

    data_structs = process_data(tickets, all_types)
    collect_metrics(auth, config["SUBDOMAIN"], data_structs["ticket_lookup"].keys(), data_structs)
    generate_excel_report(data_structs, tickets)

    elapsed = round(time.time() - start_time, 2)
    print(f"⏱ Terminé en {elapsed} secondes")
    input("Appuyez sur Entrée pour fermer...")

# ----------------------------- Lancement -----------------------------
if __name__ == "__main__":
    main()





